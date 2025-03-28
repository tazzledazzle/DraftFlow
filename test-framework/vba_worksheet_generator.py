import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def create_northshore_workbook(filename="northshore_exteriors_workbook.xlsx"):
    """
    Create a workbook with the timesheet and daily report sheets.
    This version fixes the merged cell issues.
    """
    # Create a new workbook
    wb = openpyxl.Workbook()

    # Set up the first sheet as TimeSheet
    ws_timesheet = wb.active
    ws_timesheet.title = "TimeSheet"
    create_timesheet(ws_timesheet)

    # Create the daily report sheet
    ws_daily = wb.create_sheet(title="Sheet1")
    create_daily_report(ws_daily)

    # Create other placeholder sheets seen in the tab bar
    tab_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday",
                 "Saturday", "Sunday", "Validation Table", "Job Cost Tracking"]

    for tab_name in tab_names:
        wb.create_sheet(title=tab_name)

    # Save workbook
    wb.save(filename)
    print(f"Workbook saved as: {filename}")
    return filename

def create_timesheet(ws):
    """Create the timesheet layout in the given worksheet."""
    # Set column widths
    for col_idx, width in enumerate([20, 18] + [12]*12 + [10, 10], 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Define styles
    (button_fill, center_align, gray_fill, green_fill, header_fill, header_font, left_align,
     normal_font, red_fill, subheader_font, thin_border) = define_styles()

    # Company header
    company_header(center_align, header_font, ws)

    # Foreman and Job info
    foreman_and_job_info(center_align, left_align, normal_font, subheader_font, thin_border, ws)

    # Week start info
    week_start_info(center_align, left_align, normal_font, subheader_font, thin_border, ws)

    # Days of the week header
    setup_days_of_the_week_headers(center_align, normal_font, subheader_font, thin_border, ws)

    # Now merge
    ws.merge_cells('N4:N5')

    # Add submit time button
    add_submit_time_button(button_fill, center_align, ws)

    # Hours Type headers
    ws['A5'] = "Hours Type"
    ws['A5'].font = subheader_font

    ws['B5'] = "Hours Type"
    ws['B5'].font = subheader_font

    # Worked/Sick columns
    ws['O4'] = "Worked"
    ws['O4'].font = subheader_font
    ws['O4'].border = thin_border
    ws['O4'].alignment = center_align

    ws['P4'] = "Sick"
    ws['P4'].font = subheader_font
    ws['P4'].border = thin_border
    ws['P4'].alignment = center_align

    # Add worker 1 section - MACKENZIE, DOUGAL
    current_row = 6

    # Payroll category header
    ws[f"A{current_row}"] = "PAYROLL CATEGORY"
    ws[f"A{current_row}"].font = subheader_font
    ws[f"A{current_row}"].fill = header_fill

    ws[f"B{current_row}"] = "MACKENZIE, DOUGAL"
    ws[f"B{current_row}"].font = subheader_font
    ws[f"B{current_row}"].fill = header_fill

    # Worker hours totals
    ws[f"O{current_row}"] = 17
    ws[f"O{current_row}"].font = normal_font
    ws[f"O{current_row}"].border = thin_border
    ws[f"O{current_row}"].alignment = center_align

    ws[f"P{current_row}"] = 0
    ws[f"P{current_row}"].font = normal_font
    ws[f"P{current_row}"].border = thin_border
    ws[f"P{current_row}"].alignment = center_align

    # Task rows
    tasks = [
        {"id": 1, "name": "Penthouse siding and coping", "hours": {"monday_reg": 5, "tuesday_reg": 3, "wednesday_reg": 8, "thursday_reg": 1}},
        {"id": 2, "name": "Sheet metal roof @ L14 canopy", "hours": {"tuesday_reg": 1}},
        {"id": 3, "name": "Sheet metal roof @ L14 canopy (roof build up)", "hours": {}}
    ]

    # Add task rows
    for task in tasks:
        current_row += 1

        # Task ID
        ws[f"A{current_row}"] = task["id"]
        ws[f"A{current_row}"].font = normal_font
        ws[f"A{current_row}"].border = thin_border
        ws[f"A{current_row}"].alignment = center_align

        # Task name
        ws[f"B{current_row}"] = task["name"]
        ws[f"B{current_row}"].font = normal_font
        ws[f"B{current_row}"].border = thin_border

        # Add hours
        if "monday_reg" in task["hours"]:
            ws[f"C{current_row}"] = task["hours"]["monday_reg"]
            ws[f"C{current_row}"].border = thin_border
            ws[f"C{current_row}"].alignment = center_align

        if "tuesday_reg" in task["hours"]:
            ws[f"E{current_row}"] = task["hours"]["tuesday_reg"]
            ws[f"E{current_row}"].border = thin_border
            ws[f"E{current_row}"].alignment = center_align

        if "wednesday_reg" in task["hours"]:
            ws[f"G{current_row}"] = task["hours"]["wednesday_reg"]
            ws[f"G{current_row}"].border = thin_border
            ws[f"G{current_row}"].alignment = center_align

        if "thursday_reg" in task["hours"]:
            ws[f"I{current_row}"] = task["hours"]["thursday_reg"]
            ws[f"I{current_row}"].border = thin_border
            ws[f"I{current_row}"].alignment = center_align

    # Add task buttons
    button_row = 9

    ws[f"Q{button_row}"] = "Add Task Row"
    ws[f"Q{button_row}"].font = normal_font
    ws[f"Q{button_row}"].border = thin_border
    ws[f"Q{button_row}"].alignment = center_align
    ws[f"Q{button_row}"].fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

    ws[f"Q{button_row+1}"] = "Delete Task Row"
    ws[f"Q{button_row+1}"].font = normal_font
    ws[f"Q{button_row+1}"].border = thin_border
    ws[f"Q{button_row+1}"].alignment = center_align
    ws[f"Q{button_row+1}"].fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

    # Add foreman section
    current_row += 1

    ws[f"B{current_row}"] = "Acting Foreman"
    ws[f"B{current_row}"].font = normal_font
    ws[f"B{current_row}"].border = thin_border

    current_row += 1
    ws[f"B{current_row}"] = "Sick Hours Requested"
    ws[f"B{current_row}"].font = normal_font
    ws[f"B{current_row}"].border = thin_border

    current_row += 1
    ws[f"B{current_row}"] = "Reason Off if not sick"
    ws[f"B{current_row}"].font = normal_font
    ws[f"B{current_row}"].border = thin_border

    # Add subtotals row
    current_row += 1
    for col_idx in range(3, 15):
        col_letter = get_column_letter(col_idx)
        ws[f"{col_letter}{current_row}"] = 0
        ws[f"{col_letter}{current_row}"].font = normal_font
        ws[f"{col_letter}{current_row}"].border = thin_border
        ws[f"{col_letter}{current_row}"].alignment = center_align

    # Add Worker 2 - BURTT, AARON
    current_row += 2

    # Payroll category header
    ws[f"A{current_row}"] = "PAYROLL CATEGORY"
    ws[f"A{current_row}"].font = subheader_font
    ws[f"A{current_row}"].fill = header_fill

    ws[f"B{current_row}"] = "BURTT, AARON"
    ws[f"B{current_row}"].font = subheader_font
    ws[f"B{current_row}"].fill = header_fill

    # Worker hours totals
    ws[f"O{current_row}"] = 0
    ws[f"O{current_row}"].font = normal_font
    ws[f"O{current_row}"].border = thin_border
    ws[f"O{current_row}"].alignment = center_align

    ws[f"P{current_row}"] = 0
    ws[f"P{current_row}"].font = normal_font
    ws[f"P{current_row}"].border = thin_border
    ws[f"P{current_row}"].alignment = center_align

    # Add worker 2 tasks
    for i in range(3):
        current_row += 1

        # Task ID
        ws[f"A{current_row}"] = 1
        ws[f"A{current_row}"].font = normal_font
        ws[f"A{current_row}"].border = thin_border
        ws[f"A{current_row}"].alignment = center_align

        # Task name
        if i == 0:
            ws[f"B{current_row}"] = "Column wraps @ podium"
        else:
            ws[f"B{current_row}"] = "JOB SPECIFIC TASK"

        ws[f"B{current_row}"].font = normal_font
        ws[f"B{current_row}"].border = thin_border

    # Add worker 2 foreman section
    current_row += 1

    ws[f"B{current_row}"] = "Acting Foreman"
    ws[f"B{current_row}"].font = normal_font
    ws[f"B{current_row}"].border = thin_border

    current_row += 1
    ws[f"B{current_row}"] = "Sick Hours Requested"
    ws[f"B{current_row}"].font = normal_font
    ws[f"B{current_row}"].border = thin_border

    current_row += 1
    ws[f"B{current_row}"] = "Reason Off if not sick"
    ws[f"B{current_row}"].font = normal_font
    ws[f"B{current_row}"].border = thin_border

    # Add worker 2 subtotals
    current_row += 1
    for col_idx in range(3, 15):
        col_letter = get_column_letter(col_idx)
        ws[f"{col_letter}{current_row}"] = 0
        ws[f"{col_letter}{current_row}"].font = normal_font
        ws[f"{col_letter}{current_row}"].border = thin_border
        ws[f"{col_letter}{current_row}"].alignment = center_align

    # Add Workers 3 and 4 - FULL NAME
    for worker in range(3, 5):
        current_row += 2

        # Payroll category header
        ws[f"A{current_row}"] = "PAYROLL CATEGORY"
        ws[f"A{current_row}"].font = subheader_font
        ws[f"A{current_row}"].fill = header_fill

        ws[f"B{current_row}"] = "FULL NAME"
        ws[f"B{current_row}"].font = subheader_font
        ws[f"B{current_row}"].fill = header_fill

        # Worker hours totals
        ws[f"O{current_row}"] = 0
        ws[f"O{current_row}"].font = normal_font
        ws[f"O{current_row}"].border = thin_border
        ws[f"O{current_row}"].alignment = center_align

        ws[f"P{current_row}"] = 0
        ws[f"P{current_row}"].font = normal_font
        ws[f"P{current_row}"].border = thin_border
        ws[f"P{current_row}"].alignment = center_align

        # Add tasks
        for i in range(3):
            current_row += 1

            # Task ID
            ws[f"A{current_row}"] = 1
            ws[f"A{current_row}"].font = normal_font
            ws[f"A{current_row}"].border = thin_border
            ws[f"A{current_row}"].alignment = center_align

            # Task name
            ws[f"B{current_row}"] = "JOB SPECIFIC TASK"
            ws[f"B{current_row}"].font = normal_font
            ws[f"B{current_row}"].border = thin_border

        # Add foreman section
        current_row += 1

        ws[f"B{current_row}"] = "Acting Foreman"
        ws[f"B{current_row}"].font = normal_font
        ws[f"B{current_row}"].border = thin_border

        current_row += 1
        ws[f"B{current_row}"] = "Sick Hours Requested"
        ws[f"B{current_row}"].font = normal_font
        ws[f"B{current_row}"].border = thin_border

        current_row += 1
        ws[f"B{current_row}"] = "Reason Off if not sick"
        ws[f"B{current_row}"].font = normal_font
        ws[f"B{current_row}"].border = thin_border

        # Add subtotals
        current_row += 1
        for col_idx in range(3, 15):
            col_letter = get_column_letter(col_idx)
            ws[f"{col_letter}{current_row}"] = 0
            ws[f"{col_letter}{current_row}"].font = normal_font
            ws[f"{col_letter}{current_row}"].border = thin_border
            ws[f"{col_letter}{current_row}"].alignment = center_align

    # Add summary section
    current_row += 2

    # Gray background
    for row in range(current_row, current_row + 3):
        for col in range(1, 17):
            cell = ws.cell(row=row, column=col)
            cell.fill = gray_fill

    # Worked row
    ws[f"B{current_row}"] = "Worked"
    ws[f"B{current_row}"].font = subheader_font
    ws[f"B{current_row}"].border = thin_border
    ws[f"B{current_row}"].alignment = center_align

    # Hours data
    hours_data = [5, 4, 8, 1, 0, 0, 0]
    hours_columns = ['C', 'E', 'G', 'I', 'K', 'M', 'N']

    for idx, (col, hours) in enumerate(zip(hours_columns, hours_data)):
        ws[f"{col}{current_row}"] = hours
        ws[f"{col}{current_row}"].font = normal_font
        ws[f"{col}{current_row}"].border = thin_border
        ws[f"{col}{current_row}"].alignment = center_align

    # Total
    ws[f"O{current_row}"] = 18  # Total shown in the image
    ws[f"O{current_row}"].font = subheader_font
    ws[f"O{current_row}"].border = thin_border
    ws[f"O{current_row}"].alignment = center_align

    # Sick row
    current_row += 1
    ws[f"B{current_row}"] = "Sick"
    ws[f"B{current_row}"].font = subheader_font
    ws[f"B{current_row}"].border = thin_border
    ws[f"B{current_row}"].alignment = center_align

    # Zero sick hours
    for col in hours_columns:
        ws[f"{col}{current_row}"] = 0
        ws[f"{col}{current_row}"].font = normal_font
        ws[f"{col}{current_row}"].border = thin_border
        ws[f"{col}{current_row}"].alignment = center_align

    # Total sick and extra
    ws[f"O{current_row}"] = 0
    ws[f"O{current_row}"].font = subheader_font
    ws[f"O{current_row}"].border = thin_border
    ws[f"O{current_row}"].alignment = center_align

    ws[f"P{current_row}"] = 0
    ws[f"P{current_row}"].font = normal_font
    ws[f"P{current_row}"].border = thin_border
    ws[f"P{current_row}"].alignment = center_align

    # Add/Delete Worker buttons
    current_row += 3

    # Add Worker
    ws[f"C{current_row}"] = "Add Worker +"
    ws[f"C{current_row}"].font = normal_font
    ws[f"C{current_row}"].border = thin_border
    ws[f"C{current_row}"].alignment = center_align
    ws[f"C{current_row}"].fill = green_fill

    # First create and style C, D, and E cells
    for col_letter in ['C', 'D', 'E']:
        ws[f"{col_letter}{current_row}"].fill = green_fill
        ws[f"{col_letter}{current_row}"].border = thin_border

    # Then merge
    ws.merge_cells(f'C{current_row}:E{current_row}')

    # Delete Worker
    ws[f"Q{current_row}"] = "Delete Worker"
    ws[f"Q{current_row}"].font = normal_font
    ws[f"Q{current_row}"].border = thin_border
    ws[f"Q{current_row}"].alignment = center_align
    ws[f"Q{current_row}"].fill = red_fill


def add_submit_time_button(button_fill, center_align, ws):
    ws['J1'] = "Submit Time"
    ws['J1'].font = Font(bold=True, color="FFFFFF")
    ws['J1'].fill = button_fill
    ws['J1'].alignment = center_align
    ws.merge_cells('J1:K1')


def set_ot_value(center_align, normal_font, thin_border, ws):
    # We need a different approach for merged cells
    # Create and style the cell first, with value
    cell = ws.cell(row=5, column=13)  # M5
    cell.value = "OT"
    cell.font = normal_font
    cell.border = thin_border
    cell.alignment = center_align
    # Now merge - the value will be retained in the top-left cell (M4)
    ws.merge_cells('M4:M5')


def setup_days_of_the_week_headers(center_align, normal_font, subheader_font, thin_border, ws):
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    # Monday through Friday (columns C-L)
    for i in range(5):  # 0-4 for Monday-Friday
        day_col = chr(ord('C') + i * 2)  # C, E, G, I, K

        # Set day name
        ws[f"{day_col}4"] = days[i]
        ws[f"{day_col}4"].font = subheader_font
        ws[f"{day_col}4"].border = thin_border
        ws[f"{day_col}4"].alignment = center_align

        # REG column
        ws[f"{day_col}5"] = "REG"
        ws[f"{day_col}5"].font = normal_font
        ws[f"{day_col}5"].border = thin_border
        ws[f"{day_col}5"].alignment = center_align

        # OT column
        ot_col = chr(ord(day_col) + 1)  # D, F, H, J, L
        ws[f"{ot_col}5"] = "OT"
        ws[f"{ot_col}5"].font = normal_font
        ws[f"{ot_col}5"].border = thin_border
        ws[f"{ot_col}5"].alignment = center_align
    # Saturday (column M) - set values BEFORE merging
    ws['M4'] = "Saturday"
    ws['M4'].font = subheader_font
    ws['M4'].border = thin_border
    ws['M4'].alignment = center_align

    # Set OT value
    set_ot_value(center_align, normal_font, thin_border, ws)

    # Sunday (column N) - set values BEFORE merging
    ws['N4'] = "Sunday"
    ws['N4'].font = subheader_font
    ws['N4'].border = thin_border
    ws['N4'].alignment = center_align

    # Set DT value
    cell = ws.cell(row=5, column=14)  # N5
    cell.value = "DT"
    cell.font = normal_font
    cell.border = thin_border
    cell.alignment = center_align


def week_start_info(center_align, left_align, normal_font, subheader_font, thin_border, ws):
    ws['A3'] = "Week Start Monday:"
    ws['A3'].font = subheader_font
    ws['A3'].alignment = left_align
    ws['C3'] = "3/24/25"
    ws['C3'].font = normal_font
    ws['C3'].border = thin_border
    ws['C3'].alignment = center_align
    ws['E3'] = "Project"
    ws['E3'].font = subheader_font
    ws['E3'].alignment = center_align
    ws['F3'] = 0
    ws['F3'].font = normal_font
    ws['F3'].border = thin_border
    ws['F3'].alignment = center_align
    ws['I3'] = "GC"
    ws['I3'].font = subheader_font
    ws['I3'].alignment = center_align


def foreman_and_job_info(center_align, left_align, normal_font, subheader_font, thin_border, ws):
    ws['A2'] = "Foreman:"
    ws['A2'].font = subheader_font
    ws['A2'].alignment = left_align
    ws['C2'] = "GREIN, CHRISTOPHER"
    ws['C2'].font = normal_font
    ws['C2'].border = thin_border
    ws.merge_cells('C2:D2')
    ws['E2'] = "Job #"
    ws['E2'].font = subheader_font
    ws['E2'].alignment = center_align
    ws['F2'] = 0
    ws['F2'].font = normal_font
    ws['F2'].border = thin_border
    ws['F2'].alignment = center_align
    ws['I2'] = "Shift Type"
    ws['I2'].font = subheader_font
    ws['I2'].alignment = center_align
    ws['M2'] = "Day_S-8"
    ws['M2'].font = normal_font
    ws['M2'].alignment = center_align
    ws['O2'] = "FALSE"
    ws['O2'].font = normal_font
    ws['O2'].alignment = center_align


def company_header(center_align, header_font, ws):
    ws['A1'] = "NORTHSHORE EXTERIORS"
    ws['A1'].font = header_font
    ws.merge_cells('A1:O1')
    ws['A1'].alignment = center_align


def define_styles():
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_font = Font(name='Arial', size=14, bold=True, color='000080')
    subheader_font = Font(name='Arial', size=10, bold=True)
    normal_font = Font(name='Arial', size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    header_fill = PatternFill(start_color='D0D0D0', end_color='D0D0D0', fill_type='solid')
    button_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    gray_fill = PatternFill(start_color='BBBBBB', end_color='BBBBBB', fill_type='solid')
    return button_fill, center_align, gray_fill, green_fill, header_fill, header_font, left_align, normal_font, red_fill, subheader_font, thin_border


def create_daily_report(ws):
    """Create the daily report layout."""
    # Set column widths
    for col_idx in range(1, 15):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    # Define styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_font = Font(name='Arial', size=12, bold=True)
    subheader_font = Font(name='Arial', size=10, bold=True)
    normal_font = Font(name='Arial', size=10)
    white_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')

    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    blue_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')

    # Create blue background for logo area
    for row in range(1, 8):
        ws.cell(row=row, column=1).fill = blue_fill

    # Add "NORTHSHORE" text in logo area
    ws['A5'] = "NORTHSHORE"
    ws['A5'].font = white_font
    ws['A5'].alignment = center_align

    # Company info
    ws['B2'] = "Northshore Exteriors, Inc."
    ws['B2'].font = header_font
    ws['B2'].alignment = left_align

    ws['B3'] = "11831 Beverly Park Rd, Building C"
    ws['B3'].font = normal_font
    ws['B3'].alignment = left_align

    ws['B4'] = "Everett WA 98204"
    ws['B4'].font = normal_font
    ws['B4'].alignment = left_align

    ws['B5'] = "Phone: 425.740.3700"
    ws['B5'].font = normal_font
    ws['B5'].alignment = left_align

    ws['B6'] = "Fax: 425.740.3701"
    ws['B6'].font = normal_font
    ws['B6'].alignment = left_align

    # Report title
    ws['E2'] = "Daily Report"
    ws['E2'].font = Font(name='Arial', size=16, bold=True)
    ws['E2'].alignment = center_align

    # Style and merge AFTER setting value on top-left cell
    for col in range(ord('E'), ord('I')+1):
        cell = ws.cell(row=2, column=col-64)  # Convert ASCII to column index
        cell.alignment = center_align
    ws.merge_cells('E2:I2')

    # Create Daily PDF button
    ws['L2'] = "Create Daily Pdf"
    ws['L2'].font = white_font
    ws['L2'].alignment = center_align
    ws['L2'].fill = blue_fill

    # Style and merge AFTER setting value on top-left cell
    ws.cell(row=2, column=13).fill = blue_fill  # M2
    ws.merge_cells('L2:M2')

    # Fill This Daily button
    ws['L4'] = "Fill This Daily"
    ws['L4'].font = white_font
    ws['L4'].alignment = center_align
    ws['L4'].fill = black_fill

    # Style and merge AFTER setting value on top-left cell
    ws.cell(row=4, column=13).fill = black_fill  # M4
    ws.merge_cells('L4:M4')

    # Project info boxes
    field_labels = [
        ("Project", 3, 3),
        ("Date", 4, 3),
        ("Foreman", 6, 3),
        ("Contractor", 7, 3)
    ]

    for label, row, col in field_labels:
        ws.cell(row=row, column=col).value = label
        ws.cell(row=row, column=col).font = normal_font
        ws.cell(row=row, column=col).border = thin_border

        # Add empty input box
        ws.cell(row=row, column=col+1).border = thin_border

        # If it's Date field, add Job #
        if label == "Date":
            ws.cell(row=row, column=col+2).value = "Job #"
            ws.cell(row=row, column=col+2).font = normal_font
            ws.cell(row=row, column=col+2).border = thin_border

            ws.cell(row=row, column=col+3).border = thin_border

        # If it's Foreman field, add Hrs
        if label == "Foreman":
            ws.cell(row=row, column=col+2).value = "Hrs"
            ws.cell(row=row, column=col+2).font = normal_font
            ws.cell(row=row, column=col+2).border = thin_border

    # Employees on Site section
    ws['C9'] = "Employees on Site"
    ws['C9'].font = subheader_font
    ws['C9'].alignment = center_align

    # Style and merge AFTER setting value on top-left cell
    for col in range(ord('C'), ord('I')+1):
        cell = ws.cell(row=9, column=col-64)  # Convert ASCII to column index
        cell.alignment = center_align
    ws.merge_cells('C9:I9')

    # Headers for employee table
    headers = ["Name", "Position Level", "Hrs", "Area of Site Working / Work Completed (sq/ft)"]
    col_positions = [2, 3, 4, 5]  # B, C, D, E

    for idx, header in enumerate(headers):
        col = col_positions[idx]
        cell = ws.cell(row=10, column=col)
        cell.value = header
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thin_border

    # Merge the last header
    for col in range(ord('E'), ord('I')+1):
        col_idx = col - 64  # Convert ASCII to column index
        ws.cell(row=10, column=col_idx).border = thin_border
    ws.merge_cells('E10:I10')

    # Create employee rows with borders
    for row in range(11, 26):
        for col in range(2, 10):
            ws.cell(row=row, column=col).border = thin_border

    # Add vertical area labels - IMPORTANT: First create cells, then merge

    # On Wall
    ws['E11'] = "On Wall"
    for row in range(11, 16):
        ws.cell(row=row, column=5).border = thin_border
    ws['E11'].alignment = Alignment(textRotation=90, horizontal='center', vertical='center')
    ws.merge_cells('E11:E15')

    # On Roof
    ws['E16'] = "On Roof"
    for row in range(16, 21):
        ws.cell(row=row, column=5).border = thin_border
    ws['E16'].alignment = Alignment(textRotation=90, horizontal='center', vertical='center')
    ws.merge_cells('E16:E20')

    # Other
    ws['E21'] = "Other"
    for row in range(21, 26):
        ws.cell(row=row, column=5).border = thin_border
    ws['E21'].alignment = Alignment(textRotation=90, horizontal='center', vertical='center')
    ws.merge_cells('E21:E25')

    # Quality Control Review section
    ws['B26'] = "Quality Control Review"
    ws['B26'].font = subheader_font
    ws['B26'].alignment = center_align
    ws['B26'].border = thin_border

    # Style and merge
    for col in range(ord('B'), ord('I')+1):
        col_idx = col - 64  # Convert ASCII to column index
        ws.cell(row=26, column=col_idx).border = thin_border
    ws.merge_cells('B26:I26')

    # Quality Control rows
    for row in range(27, 30):
        for col in range(2, 10):
            ws.cell(row=row, column=col).border = thin_border

    # Signature section
    ws['C30'] = "Date:"
    ws['C30'].font = normal_font
    ws['C30'].alignment = right_align

    ws['D30'] = "3/27/25"
    ws['D30'].font = normal_font
    ws['D30'].alignment = center_align
    ws['D30'].border = thin_border

    ws['G30'] = "Printed Name"
    ws['G30'].font = normal_font
    ws['G30'].alignment = right_align

    ws['H30'] = "0"
    ws['H30'].font = normal_font
    ws['H30'].alignment = center_align
    ws['H30'].border = thin_border

    # Terms section
    ws['B32'] = "The above and beyond that which are required per the contract documents, and Northshore Exteriors, Inc. shall be compensated for above overtime hours for the premium which Northshore is required to pay our pliice increment and profit at 22%"
    ws['B32'].font = normal_font
    ws['B32'].alignment = left_align

    # Style and merge
    for col in range(ord('B'), ord('L')+1):
        col_idx = col - 64  # Convert ASCII to column index
        cell = ws.cell(row=32, column=col_idx)
        cell.alignment = left_align
    ws.merge_cells('B32:L32')

    # Equipment on Site section
    ws['C34'] = "Equipment on Site"
    ws['C34'].font = subheader_font
    ws['C34'].alignment = center_align

    # Style and merge
    for col in range(ord('C'), ord('E')+1):
        col_idx = col - 64  # Convert ASCII to column index
        cell = ws.cell(row=34, column=col_idx)
        cell.alignment = center_align
    ws.merge_cells('C34:E34')

    # Equipment headers
    equip_headers = ["Own/Rent", "Equipment/Name", "Size", "Hours", "Notes"]
    for i, header in enumerate(equip_headers):
        col = i + 2  # Start at column B
        ws.cell(row=35, column=col).value = header
        ws.cell(row=35, column=col).font = normal_font
        ws.cell(row=35, column=col).alignment = center_align
        ws.cell(row=35, column=col).border = thin_border

    # Equipment rows
    for row in range(36, 40):
        for col in range(2, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if col == 2:
                cell.value = "O/R"

    # Major Milestones section
    ws['I34'] = "Major Milestones"
    ws['I34'].font = subheader_font
    ws['I34'].alignment = center_align

    # Style and merge
    for col in range(ord('I'), ord('K')+1):
        col_idx = col - 64  # Convert ASCII to column index
        cell = ws.cell(row=34, column=col_idx)
        cell.alignment = center_align
    ws.merge_cells('I34:K34')

    # Milestones box
    for row in range(35, 37):
        for col in range(9, 12):
            ws.cell(row=row, column=col).border = thin_border

    # Changes Directed section
    ws['I38'] = "Changes Directed or Proposed"
    ws['I38'].font = subheader_font
    ws['I38'].alignment = center_align

    # Style and merge
    for col in range(ord('I'), ord('K')+1):
        col_idx = col - 64  # Convert ASCII to column index
        cell = ws.cell(row=38, column=col_idx)
        cell.alignment = center_align
    ws.merge_cells('I38:K38')

    # Changes box
    for row in range(39, 41):
        for col in range(9, 12):
            ws.cell(row=row, column=col).border = thin_border

    # Meetings/Visitors section
    ws['C40'] = "Meetings / Visitors"
    ws['C40'].font = subheader_font
    ws['C40'].alignment = center_align

    # Style and merge
    for col in range(ord('C'), ord('E')+1):
        col_idx = col - 64  # Convert ASCII to column index
        cell = ws.cell(row=40, column=col_idx)
        cell.alignment = center_align
    ws.merge_cells('C40:E40')

    # Meeting headers
    meeting_headers = ["Time", "Attendees", "Topics"]

    # Column B - Time
    ws['B41'] = meeting_headers[0]
    ws['B41'].font = normal_font
    ws['B41'].alignment = center_align
    ws['B41'].border = thin_border

    # Column C - Attendees
    ws['C41'] = meeting_headers[1]
    ws['C41'].font = normal_font
    ws['C41'].alignment = center_align
    ws['C41'].border = thin_border

    # Columns D-F - Topics
    ws['D41'] = meeting_headers[2]
    ws['D41'].font = normal_font
    ws['D41'].alignment = center_align
    ws['D41'].border = thin_border

    # Style and merge
    for col in range(ord('D'), ord('F')+1):
        col_idx = col - 64  # Convert ASCII to column index
        cell = ws.cell(row=41, column=col_idx)
        cell.border = thin_border
        cell.alignment = center_align
    ws.merge_cells('D41:F41')

    # Meeting rows
    for row in range(42, 45):
        for col in range(2, 7):
            ws.cell(row=row, column=col).border = thin_border

    # Deliveries section
    ws['C46'] = "Deliveries"
    ws['C46'].font = subheader_font
    ws['C46'].alignment = center_align

    # Style and merge
    for col in range(ord('C'), ord('E')+1):
        col_idx = col - 64  # Convert ASCII to column index
        cell = ws.cell(row=46, column=col_idx)
        cell.alignment = center_align
    ws.merge_cells('C46:E46')

    # Delivery headers
    delivery_headers = ["From", "Materials Included", "Placing Site"]
    for i, header in enumerate(delivery_headers):
        col = i + 2  # Start at column B
        ws.cell(row=47, column=col).value = header
        ws.cell(row=47, column=col).font = normal_font
        ws.cell(row=47, column=col).alignment = center_align
        ws.cell(row=47, column=col).border = thin_border

    # Delivery rows
    for row in range(48, 52):
        for col in range(2, 5):
            ws.cell(row=row, column=col).border = thin_border

    # Accidents section
    ws['I46'] = "Accidents, Incidents, Hazards, or Safety Roomnotes"
    ws['I46'].font = subheader_font
    ws['I46'].alignment = center_align

    # Style and merge
    for col in range(ord('I'), ord('K')+1):
        col_idx = col - 64  # Convert ASCII to column index
        cell = ws.cell(row=46, column=col_idx)
        cell.alignment = center_align
    ws.merge_cells('I46:K46')

    # Accidents box
    for row in range(47, 52):
        for col in range(9, 12):
            ws.cell(row=row, column=col).border = thin_border


if __name__ == "__main__":
    file_path = create_northshore_workbook()